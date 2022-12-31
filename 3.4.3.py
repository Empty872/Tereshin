import csv
import os
import sys
from operator import itemgetter
import pandas as pd
import math
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Report:
    '''
    Класс для создания отчетов
    Attributes:
        inputValues (InputConect) входные данные
        yearSalary (dict) зп по годам
        yearSalary_needed (dict) зп по  годам для вакансии
        year_to_count (dict) ваканси по годам
        yearCount (dict) вакансии по годам для профессии
        areaSalary (dict) уровень зп по городам
        areaPiece (dict) доля вакансий по городам
    '''

    def __init__(self, inputValues, yearSalary, yearSalary_needed, year_to_count, yearCount, areaSalary, areaPiece):
        self.inputValues = inputValues
        self.yearSalary = yearSalary
        self.yearSalary_needed = yearSalary_needed
        self.year_to_count = year_to_count
        self.yearCount = yearCount
        self.areaSalary = areaSalary
        self.areaPiece = areaPiece
        self.generate_pdf()

    def generate_excel(self):
        '''
        Создает файл отчета в экселе
        :return: void
        '''
        excelFile = Workbook()
        excelSheetFirst = excelFile.create_sheet(title="Статистика по годам", index=0)
        excelSheetFirst['A1'] = 'Год'
        excelSheetFirst['B1'] = 'Средняя зарплата'
        excelSheetFirst['C1'] = f'Средняя зарплата - {self.inputValues.professionName}, {self.inputValues.areaName}'
        excelSheetFirst['D1'] = 'Количество вакансий'
        excelSheetFirst['E1'] = f'Количество вакансий - {self.inputValues.professionName}, {self.inputValues.areaName}'
        excelSheetFirst['A1'].font = Font(bold=True)
        excelSheetFirst['B1'].font = Font(bold=True)
        excelSheetFirst['C1'].font = Font(bold=True)
        excelSheetFirst['D1'].font = Font(bold=True)
        excelSheetFirst['E1'].font = Font(bold=True)
        yearRow = list(self.yearSalary.keys())
        firstValue = list(self.yearSalary.values())
        secondValues = list(self.yearSalary_needed.values())
        thirdValues = list(self.year_to_count.values())
        fourthValues = list(self.yearCount.values())
        for i in range(0, 16):
            data = list()
            data.append(yearRow[i])
            data.append(firstValue[i])
            data.append(secondValues[i])
            data.append(thirdValues[i])
            data.append(fourthValues[i])
            excelSheetFirst.append(data)

        self.setBorder(columns=['A', 'B', 'C', 'D', 'E'], excelSheet=excelSheetFirst, numberSheet=0)
        self.setSizeColumns(excelSheet=excelSheetFirst)

        excelSheetSecond = excelFile.create_sheet(title="Статистика по городам", index=1)

        excelSheetSecond['A1'] = 'Город'
        excelSheetSecond['B1'] = 'Уровень зарплат'
        excelSheetSecond['D1'] = 'Город'
        excelSheetSecond['E1'] = 'Доля вакансий'
        excelSheetSecond['A1'].font = Font(bold=True)
        excelSheetSecond['B1'].font = Font(bold=True)
        excelSheetSecond['D1'].font = Font(bold=True)
        excelSheetSecond['E1'].font = Font(bold=True)
        cityes1 = list(self.areaSalary.keys())
        salaryes = list(self.areaSalary.values())
        cityes2 = list(self.areaPiece.keys())
        pieces = list(self.areaPiece.values())
        for i in range(0, 10):
            data = list()
            data.append(cityes1[i])
            data.append(salaryes[i])
            data.append("")
            data.append(cityes2[i])
            data.append(pieces[i])
            excelSheetSecond.append(data)

        self.setBorder(columns=['A', 'B', 'D', 'E'], excelSheet=excelSheetSecond, numberSheet=1)
        self.setSizeColumns(excelSheet=excelSheetSecond)

        excelFile.save('report.xlsx')

    def setSizeColumns(self, excelSheet):
        '''
        Устанавливает размер колонок
        :param excelSheet: инстанс листа эксель
        :return: void
        '''
        i = 0
        col_width = list()
        for col in excelSheet.columns:
            for j in range(len(col)):
                if j == 0:
                    col_width.append(len(str(col[j].value)))
                else:
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1

        for i in range(len(col_width)):
            col_letter = get_column_letter(i + 1)
            if col_width[i] > 100:
                excelSheet.column_dimensions[col_letter].width = 100
            else:
                excelSheet.column_dimensions[col_letter].width = col_width[i] + 2

    def setBorder(self, columns, excelSheet, numberSheet):
        '''
        Устанавливает рамки в экселе
        :param columns: колонка в экселе
        :param excelSheet: инстанс листа в экселе
        :param numberSheet: номер листа
        :return: void
        '''
        side = Side(border_style='thin', color="00000000")
        for i in columns:
            column = excelSheet[i]
            for j in column:
                j.border = Border(left=side, right=side, top=side, bottom=side)
                if i == 'E' and numberSheet == 1:
                    j.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

    def generate_pdf(self):
        '''
        Генерирует пдф отчет
        :return: void
        '''
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template_3_4_3.html')
        heads1 = ['Год', 'Средняя зарплата',
                  f'Средняя зарплата - {self.inputValues.professionName}, {self.inputValues.areaName}',
                  'Количество вакансий',
                  f'Количество вакансий - {self.inputValues.professionName}, {self.inputValues.areaName}']
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        for key in self.areaPiece.keys():
            self.areaPiece[key] = str(round(self.areaPiece[key] * 100, 2)) + '%'
        pdf_template = template.render({"yearSalary": self.yearSalary,
                                        "yearSalary_needed": self.yearSalary_needed,
                                        "year_to_count": self.year_to_count,
                                        "yearCount": self.yearCount,
                                        "areaSalary": self.areaSalary,
                                        "areaPiece": self.areaPiece,
                                        "heads1": heads1,
                                        "heads2": heads2,
                                        "profession": self.inputValues.professionName,
                                        "areaName": self.inputValues.areaName})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report_3_4_3.pdf', configuration=config,
                           options={"enable-local-file-access": None})


class DataSet:
    def __init__(self):
        self.folder_name = 'very_new_csv_files'
        self.inputValues = InputConect()
        self.vacancy = self.inputValues.professionName
        self.areaName = self.inputValues.areaName
        self.df = pd.read_csv(self.inputValues.fileName)

        self.df['salary_mean'] = self.df[['salary_from', 'salary_to']].mean(axis=1)
        self.df['salary'] = self.df.apply(GetSalaryToRub, axis=1)
        self.df['published_at'] = self.df['published_at'].apply(lambda x: int(x[:4]))
        self.years = self.df['published_at'].unique()
        self.salaryByYears = {year: [] for year in self.years}
        self.vacsByYears = {year: 0 for year in self.years}
        self.vacSalaryByYears = {year: [] for year in self.years}
        self.vacCountByYears = {year: 0 for year in self.years}
        self.GetStaticByCities()
        self.initializeYearStatistics()

    def initializeYearStatistics(self):
        """Добавляет в словари статистик значения из файла
            Также запускает concurrent.futures
        """
        for i in range(2003, 2023):
            result = self.getStatisticByYear(i)
            print(result)
            self.salaryByYears[i] = result[1]
            self.vacsByYears[i] = result[2]
            self.vacSalaryByYears[i] = result[3]
            self.vacCountByYears[i] = result[4]
        self.salaryByYears = dict(sorted(self.salaryByYears.items()))
        self.vacsByYears = dict(sorted(self.vacsByYears.items()))
        self.vacSalaryByYears = dict(sorted(self.vacSalaryByYears.items()))
        self.vacCountByYears = dict(sorted(self.vacCountByYears.items()))
        self.PrintInfo()

    def getStatisticByYear(self, year):
        """Возвращает статистку за год в порядке:
            Среднее значение зарплаты за год,
            Количество вакансий за год.
            Среднее значение зарплаты за год для выбранной профессии.
            Количество вакансий за год для выбранной профессии
        """
        file_path = rf"{self.folder_name}/new_csv_{year}.csv"
        if os.path.exists(file_path):
            df = pd.read_csv(file_path, on_bad_lines='skip')
            df["salary_mean"] = df[["salary_from", "salary_to"]].mean(axis=1)
            df['salary'] = df.apply(GetSalaryToRub, axis=1)
            df_vacancy = df[df["name"].str.contains(self.vacancy)]
            df_vacancy = df_vacancy[df_vacancy["area_name"].str.contains(self.areaName)]
            averageSalary = math.floor(df["salary"].mean())
            numberOfVacancies = len(df.index)
            averageSalaryProfession = 0 if df_vacancy.empty else math.floor(df_vacancy["salary"].mean())
            numberOfVacanciesProfession = 0 if df_vacancy.empty else len(df_vacancy.index)

            return [year, averageSalary, numberOfVacancies, averageSalaryProfession,
                    numberOfVacanciesProfession]

    def PrintInfo(self):
        '''
        Вывод данных
        :return: void
        '''
        print('Динамика уровня зарплат по годам:', self.salaryByYears)
        print('Динамика количества вакансий по годам:', self.vacsByYears)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.vacSalaryByYears)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.vacCountByYears)
        print('Уровень зарплат по городам (в порядке убывания):', self.salaryArea)
        print('Доля вакансий по городам (в порядке убывания):', self.countArea)
        Report(inputValues=self.inputValues, yearSalary=self.salaryByYears, yearSalary_needed=self.vacSalaryByYears,
               year_to_count=self.vacsByYears, yearCount=self.vacCountByYears, areaSalary=self.salaryArea,
               areaPiece=self.countArea)

    def GetStaticByCities(self):
        '''
        Получение данных по городам, не требующих multiprocessing
        :return: void
        '''
        total = len(self.df)
        self.df['count'] = self.df.groupby('area_name')['area_name'].transform('count')
        # округление 1% по условию задачи
        df_norm = self.df[self.df['count'] > 0.01 * total]
        df_area = df_norm.groupby('area_name', as_index=False)['salary'].mean().sort_values(by='salary',
                                                                                            ascending=False)
        df_count = df_norm.groupby('area_name', as_index=False)['count'].mean().sort_values(by='count', ascending=False)
        cities = df_count['area_name'].unique()

        self.salaryArea = {city: 0 for city in cities}
        self.countArea = {city: 0 for city in cities}
        for city in cities:
            self.salaryArea[city] = int(df_area[df_area['area_name'] == city]['salary'])
            self.salaryArea = dict(sorted(self.salaryArea.items(), key=lambda x: x[1], reverse=True))
            self.countArea[city] = round(int(df_count[df_count['area_name'] == city]['count']) / total, 4)
        salaryArea = {}
        countArea = {}
        sorted_salaryArea = dict(sorted(self.salaryArea.items(), key=itemgetter(1), reverse=True))
        sorted_countArea = dict(sorted(self.countArea.items(), key=itemgetter(1), reverse=True))
        for key in [key for key in sorted_salaryArea][:10]:
            salaryArea[key] = self.salaryArea[key]
        for key in [key for key in sorted_countArea][:10]:
            countArea[key] = self.countArea[key]
        self.salaryArea = salaryArea
        self.countArea = countArea


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


def GetSalaryToRub(row):
    if math.isnan(row['salary_mean']):
        return 0
    return row['salary_mean'] * currency_to_rub[row['salary_currency']]


class InputConect:
    '''
    Класс для получения входных данных и их валидации
    Attributes:
        fileName (string) имя файла для получения статистики
        professionName (string) название профессии, по которой нужна статистика
    '''

    def __init__(self):
        self.fileName = input("Введите название файла: ")
        self.professionName = input('Введите название профессии: ')
        self.areaName = input('Введите название города: ')
        self.checkFile()

    def checkFile(self):
        '''
        Валидация входных полей
        :return: void
        '''
        with open(self.fileName, "r", encoding='utf-8-sig', newline='') as csv_file:
            file_iter = iter(csv.reader(csv_file))
            if next(file_iter, "none") == "none":
                print("Пустой файл")
                sys.exit()

            if next(file_iter, "none") == "none":
                print("Нет данных")
                sys.exit()


DataSet()
